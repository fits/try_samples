
import sys

from statistics import mean
from toolz import concat, frequencies

from gensim.corpora import Dictionary
from gensim.models.ldamodel import LdaModel
from gensim.models import word2vec
from openpyxl import Workbook

data_file = sys.argv[1]
topic_num = int(sys.argv[2])
alpha = float(sys.argv[3])
dest_file = sys.argv[4]

sentences = list(word2vec.LineSentence(data_file))

dic = Dictionary(sentences)

corpus = [dic.doc2bow(s) for s in sentences if len(s) >= 2]

lda = LdaModel(corpus = corpus, id2word = dic, num_topics = topic_num, alpha = alpha, random_state = 1)

doc_topics = [lda[c] for c in corpus]

avg_doc_topics = mean([len(t) for t in doc_topics])

print(f"topics num of doc = {avg_doc_topics}")

topic_freq = frequencies(concat([[x[0] for x in t] for t in doc_topics]))

wb = Workbook()

sh1 = wb.active
sh1.title = 'topics'

sh1.append(['topic', 'freq', 'item', 'prob'])

for i in range(topic_num):
  for t in lda.get_topic_terms(i):
    item = dic[t[0]]

    sh1.append([i, topic_freq[i], item, t[1]])

sh1.auto_filter.ref = f"A1:D{sh1.max_row}"

sh2 = wb.create_sheet('corpus')

sh2.append(['corpus', 'item', 'topic'])

for i in range(len(corpus)):
  dts = lda.get_document_topics(corpus[i], per_word_topics = True)

  for dt in dts[1]:
    item = dic[dt[0]]
    sh2.append([i, item, dt[1][0]])

sh2.auto_filter.ref = f"A1:C{sh2.max_row}"

wb.save(dest_file)
